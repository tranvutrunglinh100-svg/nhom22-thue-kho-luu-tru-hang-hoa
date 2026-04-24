# ============================================================
# MODULE 7: TÀI CHÍNH & THANH TOÁN (Billing)
# File: app/routes/billing.py
# ============================================================

from flask import (Blueprint, render_template, redirect, url_for,
                   flash, request, jsonify)
from flask_login import login_required, current_user
from datetime import date, timedelta
from sqlalchemy import func
from app import db
from app.models import (Invoice, Payment, InvoiceStatus, Contract,
                        ContractStatus, Customer)

billing_bp = Blueprint('billing', __name__, url_prefix='/billing')

STATUS_LABELS = {
    InvoiceStatus.UNPAID.value:  ('Còn nợ',         'warning'),
    InvoiceStatus.PAID.value:    ('Đã thanh toán',   'success'),
    InvoiceStatus.OVERDUE.value: ('Quá hạn',         'danger'),
}


def _generate_invoice_code():
    last = Invoice.query.order_by(Invoice.id.desc()).first()
    seq  = (last.id + 1) if last else 1
    return f'INV-{seq:04d}'


# ─── Dashboard tài chính ─────────────────────────────────

@billing_bp.route('/')
@login_required
def index():
    if not current_user.is_admin():
        flash('Chỉ Admin mới có quyền truy cập tài chính.', 'danger')
        return redirect(url_for('dashboard.index'))

    # Cập nhật hóa đơn quá hạn
    overdue_invoices = Invoice.query.filter(
        Invoice.status == InvoiceStatus.UNPAID.value,
        Invoice.due_date < date.today()
    ).all()
    for inv in overdue_invoices:
        inv.status = InvoiceStatus.OVERDUE.value
    db.session.commit()

    status_filter = request.args.get('status', 'all')
    search        = request.args.get('search', '').strip()
    page          = request.args.get('page', 1, type=int)

    query = Invoice.query.join(Contract).join(Customer)

    if status_filter != 'all':
        query = query.filter(Invoice.status == status_filter)
    if search:
        query = query.filter(Customer.name.ilike(f'%{search}%') |
                             Invoice.code.ilike(f'%{search}%'))

    invoices = query.order_by(Invoice.due_date).paginate(
        page=page, per_page=15, error_out=False
    )

    # Thống kê tổng
    first_day = date.today().replace(day=1)
    stats = {
        'paid_month': db.session.query(func.sum(Invoice.total_amount)).filter(
            Invoice.status == InvoiceStatus.PAID.value,
            Invoice.paid_date >= first_day
        ).scalar() or 0,
        'unpaid': db.session.query(func.sum(Invoice.total_amount)).filter(
            Invoice.status == InvoiceStatus.UNPAID.value
        ).scalar() or 0,
        'overdue': db.session.query(func.sum(Invoice.total_amount)).filter(
            Invoice.status == InvoiceStatus.OVERDUE.value
        ).scalar() or 0,
        'total_invoices': Invoice.query.count(),
    }

    return render_template('billing/index.html',
                           invoices=invoices,
                           status_filter=status_filter,
                           search=search,
                           stats=stats,
                           STATUS_LABELS=STATUS_LABELS,
                           InvoiceStatus=InvoiceStatus)


# ─── Chi tiết hóa đơn ────────────────────────────────────

@billing_bp.route('/<int:invoice_id>')
@login_required
def detail(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    invoice.check_overdue()
    db.session.commit()
    return render_template('billing/detail.html',
                           invoice=invoice,
                           STATUS_LABELS=STATUS_LABELS)


# ─── Tạo hóa đơn thủ công ────────────────────────────────

@billing_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create():
    if not current_user.is_admin():
        flash('Chỉ Admin mới có quyền tạo hóa đơn.', 'danger')
        return redirect(url_for('billing.index'))

    contracts = Contract.query.filter(
        Contract.status.in_([ContractStatus.ACTIVE.value,
                              ContractStatus.EXPIRING.value])
    ).all()

    if request.method == 'POST':
        contract_id    = int(request.form.get('contract_id'))
        period_start   = date.fromisoformat(request.form.get('period_start'))
        period_end     = date.fromisoformat(request.form.get('period_end'))
        service_amount = float(request.form.get('service_amount', 0) or 0)
        due_date_str   = request.form.get('due_date')
        notes          = request.form.get('notes', '').strip()

        contract = Contract.query.get(contract_id)

        # Tính tiền thuê theo số ngày
        days        = (period_end - period_start).days + 1
        daily_rate  = contract.monthly_rate / 30
        base_amount = round(daily_rate * days, 0)
        total_amount = base_amount + service_amount

        invoice = Invoice(
            code           = _generate_invoice_code(),
            contract_id    = contract_id,
            period_start   = period_start,
            period_end     = period_end,
            base_amount    = base_amount,
            service_amount = service_amount,
            total_amount   = total_amount,
            due_date       = date.fromisoformat(due_date_str) if due_date_str else
                             period_end + timedelta(days=15),
            notes          = notes
        )
        db.session.add(invoice)
        db.session.commit()
        flash(f'Tạo hóa đơn {invoice.code} thành công!', 'success')
        return redirect(url_for('billing.detail', invoice_id=invoice.id))

    return render_template('billing/form.html',
                           invoice=None, contracts=contracts)


# ─── Tự động tạo hóa đơn hàng tháng ─────────────────────

@billing_bp.route('/auto-generate', methods=['POST'])
@login_required
def auto_generate():
    """Tạo hóa đơn tự động cho tất cả hợp đồng đang hoạt động"""
    if not current_user.is_admin():
        return jsonify({'error': 'Không có quyền'}), 403

    today      = date.today()
    first_day  = today.replace(day=1)
    if today.month == 12:
        last_day = date(today.year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(today.year, today.month + 1, 1) - timedelta(days=1)

    contracts = Contract.query.filter(
        Contract.status.in_([ContractStatus.ACTIVE.value,
                              ContractStatus.EXPIRING.value])
    ).all()

    created_count = 0
    for contract in contracts:
        # Kiểm tra đã có hóa đơn kỳ này chưa
        existing = Invoice.query.filter_by(
            contract_id  = contract.id,
            period_start = first_day,
            period_end   = last_day
        ).first()

        if not existing:
            invoice = Invoice(
                code           = _generate_invoice_code(),
                contract_id    = contract.id,
                period_start   = first_day,
                period_end     = last_day,
                base_amount    = contract.monthly_rate,
                service_amount = 0,
                total_amount   = contract.monthly_rate,
                due_date       = last_day + timedelta(days=15),
            )
            db.session.add(invoice)
            created_count += 1

    db.session.commit()
    flash(f'Đã tạo {created_count} hóa đơn tháng {today.month}/{today.year}.', 'success')
    return redirect(url_for('billing.index'))


# ─── Ghi nhận thanh toán ─────────────────────────────────

@billing_bp.route('/<int:invoice_id>/pay', methods=['GET', 'POST'])
@login_required
def record_payment(invoice_id):
    if not current_user.is_admin():
        flash('Chỉ Admin mới có quyền ghi nhận thanh toán.', 'danger')
        return redirect(url_for('billing.index'))

    invoice = Invoice.query.get_or_404(invoice_id)

    if request.method == 'POST':
        amount    = float(request.form.get('amount', 0))
        method    = request.form.get('method', 'Chuyển khoản')
        reference = request.form.get('reference', '').strip()

        payment = Payment(
            invoice_id  = invoice.id,
            amount      = amount,
            method      = method,
            reference   = reference,
            recorded_by = current_user.id
        )
        db.session.add(payment)

        # Tổng thanh toán
        total_paid = sum(p.amount for p in invoice.payments) + amount
        if total_paid >= invoice.total_amount:
            invoice.status    = InvoiceStatus.PAID.value
            invoice.paid_date = date.today()
        else:
            invoice.status = InvoiceStatus.UNPAID.value

        db.session.commit()
        flash(f'Ghi nhận thanh toán {amount:,.0f}đ thành công!', 'success')
        return redirect(url_for('billing.detail', invoice_id=invoice_id))

    return render_template('billing/payment_form.html', invoice=invoice)


# ─── Xuất Excel danh sách hóa đơn ────────────────────────

@billing_bp.route('/export')
@login_required
def export_excel():
    """Xuất danh sách hóa đơn ra Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    from flask import send_file

    invoices = Invoice.query.join(Contract).join(Customer).order_by(
        Invoice.due_date.desc()
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Danh sách hóa đơn'

    # Header
    headers = ['Mã HĐ', 'Khách hàng', 'Kỳ', 'Tiền thuê (đ)',
               'Tiền dịch vụ (đ)', 'Tổng tiền (đ)', 'Hạn TT', 'Trạng thái']
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill(fill_type='solid', fgColor='1F6EA0')
        cell.alignment = Alignment(horizontal='center')

    # Dữ liệu
    STATUS_VI = {
        'paid': 'Đã thanh toán',
        'unpaid': 'Còn nợ',
        'overdue': 'Quá hạn'
    }
    for inv in invoices:
        ws.append([
            inv.code,
            inv.contract.customer.name,
            f'{inv.period_start.strftime("%d/%m/%Y")} - {inv.period_end.strftime("%d/%m/%Y")}',
            inv.base_amount,
            inv.service_amount,
            inv.total_amount,
            inv.due_date.strftime('%d/%m/%Y') if inv.due_date else '',
            STATUS_VI.get(inv.status, inv.status)
        ])

    # Định dạng cột
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 28

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'hoa_don_{date.today().strftime("%Y%m%d")}.xlsx'
    )
